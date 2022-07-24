# -*- coding: utf-8 -*-

import telebot
from openpyxl import load_workbook
import pandas
import dataframe_image
import datetime
from io import BytesIO
import os
import time
import configparser
from sys import getdefaultencoding
import random
import threading
import schedule
import sys
import jira
import imgkit
from jira.client import JIRA
from datetime import date
import urllib.request
import json
import requests
import re
import base64
import pyodbc

lastseen = None
nightqueue = 0
address = r'C:\...path to .xlsx for Duty function'

#procedure for loading excel file and parsing its data into dict:
def ParseData(file_addr) -> dict:
  duty_data = load_workbook(file_addr)

  duty_data_stored = list()
  for rows in duty_data["Дежурства"].values:
      duty_data_stored.append([x for x in rows])

  duty_data_final = dict()
  #creating dictionary of lists with dict key based on first element in column:
  for col in range(len(duty_data_stored[0])):
      for i in range(len(duty_data_stored)):
          if i == 0:
              duty_data_final[duty_data_stored[i][col]] = list()
          else:
              if type (duty_data_stored[i][col]) == int: duty_data_stored[i][col] = str(duty_data_stored[i][col]) #this is for phone numbers
              duty_data_final[duty_data_stored[0][col]].append(duty_data_stored[i][col])

  return duty_data_final

#conditionally remove stuff. accepts dept, name and time params:
def Conditions(duty, **kwargs):
    empty = 0
    while empty < len(duty['ФИО']):
        if duty['ФИО'][empty] == None:
            for x in duty.keys(): del duty[x][empty]
        else:
            empty+=1
    if len(kwargs) == 0:
        Logging(0, "for some reason empty conditions call.")
        return
    period = kwargs['range']
    if 'dept' in kwargs and len(kwargs['dept']) > 0:
        i = 0
        while i < len(duty['Группа']):
            if duty['Группа'][i] != kwargs['dept']:
                for x in duty.keys(): del duty[x][i]
            else:
                i+=1
    if 'name' in kwargs:
        i = 0
        while i < len(duty['ФИО']):
            namebool = True
            for name in kwargs['name']:
                if name.lower() not in duty['ФИО'][i].lower(): namebool = False
            if not namebool:
                for x in duty.keys(): del duty[x][i]
            else:
                i+=1
    if 'time' in kwargs:
        i = 0
        while i < len(duty['С']):
            if duty['По'][i] < kwargs['time'] or duty['С'][i] > kwargs['time']+datetime.timedelta(days=period):
                for x in duty.keys(): del duty[x][i]
            else:
                i+=1
    return

#main procedure for generating duty table:
def GetData(file_addr, group, namecheck, customrange) -> BytesIO:
   if customrange == -1: customrange = 7 #default range. -1 in case it isn't specified
   duty = ParseData(file_addr)
   Logging(0, 'Loaded duty table and preparing for filering')
   if len(namecheck) > 0:
       Conditions(duty, dept=group, time=datetime.datetime.now(), name = namecheck, range = customrange)
   else:
       Conditions(duty, dept=group, time=datetime.datetime.now(), range = customrange)
   if len(duty['ФИО']) == 0: duty = dict({'Ничего':[], 'не':[], 'нашел.':[], 'Извините':[]})
   else: duty = {duty_res: v for duty_res, v in duty.items() if duty_res in {'ФИО', 'С', 'По', 'Номер'}}
   hello = pandas.DataFrame(data = duty)
   Logging(0, 'Loaded filtered table in DataFrame object')
   img = BytesIO()
   img.name = 'duty.png'
   hello.index = hello.index + 1
   dataframe_image.export(hello, img)
   Logging(0, 'Exported table into image')
   return img

def DutyInform(theBot, group, groupchatid, chatid): #not used.
    for thread in threading.enumerate():
        print(thread.name)
    Logging(0, "Preparing a scheduled message.")
    Logging(0, 'Preparing duty table')
    if len(group) > 0:
       grouptext = 'группе ' + group
    else:
       grouptext = 'всем группам'
    try:
       img = GetData(address, group, '', -1)
    except Exception as exception:
       Logging(1, 'Exception occured: ' + str(exception))
       #I dont think I need this during weekly informs:
       #theBot.send_message(groupchatid, "Не смог получить график. Либо не найдена нужная таблица, либо данных нет или они в некорректном формате.")
       time.sleep(60)
    else:
       img.seek(0)
       theBot.send_message(groupchatid, "Автоматическое еженедельное уведомление. График дежурств по " + grouptext + " на ближайшую неделю: ")
       theBot.send_photo(groupchatid, photo=img)
       Logging(0, 'Sent duty table')
       time.sleep(60)
    return

#small function for useful links
def UsefulLinksList(file_addr) -> list:
    links = load_workbook(file_addr)
    linksdict = list()
    for rows in links["Links"].values:
        linksdict.append([x for x in rows])
    return linksdict

#gets alert from jira
def GetAlert(theBot, chatid):
    global lastseen
    global nightqueue
    with open(os.path.dirname(os.path.abspath(__file__)) + '\\forjira.txt', 'r') as f:
        creds = f.read().split(" ")

    jira_unit = JIRA(basic_auth=(creds[0], creds[1]), options={'server': 'our jira server'})
    
    for issue in jira_unit.search_issues(
            'project = our_project AND \"Epic link\" = \"our epic\" AND component = \"our component\" order by created desc',
            maxResults=5):
        print(issue.key) #store this
        if nightqueue > 0 and not(datetime.time(0) <= datetime.datetime.now().time() <= datetime.time(9)):
                night = open(os.path.dirname(os.path.abspath(__file__)) + '\\nightqueue\\night.txt', 'r')
                nightlist = night.read().split(" ")
                if len(nightlist[-1]) == 0: del nightlist[-1] #crude way of dealing with empty element after split
                #print(nightlist)
                theBot.send_message(chatid, "Доброе утро! Количество алертов за ночь: "+str(len(nightlist))+".")
                for iss in nightlist:
                    theBot.send_message(chatid, "our jira server"+iss)
                    try:
                        theBot.send_photo(chatid, photo=open(os.path.dirname(os.path.abspath(__file__))+'\\nightqueue\\'+iss+'.png', 'rb'))
                    except Exception as e:
                        if "PHOTO_INVALID_DIMENSIONS" in str(e):
                            Logging(1,
                                    "Could not send photo for alert " + iss + " using send_photo method. PHOTO_INVALID_DIMENSIONS encountered. Using send_document instead.")
                            theBot.send_document(chatid, open(os.path.dirname(os.path.abspath(__file__))+'\\nightqueue\\'+iss+'.png', 'rb'))
                        else:
                            Logging(1, "Exception: " + str(e) + " when sending photo for alert " + iss)
                    os.remove(os.path.dirname(os.path.abspath(__file__))+'\\nightqueue\\'+iss+'.png')
                nightqueue = 0
                open(os.path.dirname(os.path.abspath(__file__)) + '\\nightqueue\\night.txt', 'w').close()
        if issue.key not in lastseen:
            if not(datetime.time(0) <= datetime.datetime.now().time() <= datetime.time(9)): #and not(datetime.datetime.now().weekday() < 5 and datetime.time(9) <= datetime.datetime.now().time() <= datetime.time(18)) and lastseen != None: #!=None for init procedure
                linky = 'our jira server' + issue.key
                img = imgkit.from_string("<meta charset=\"utf-8\">\n" + issue.fields.description, False,
                                         config=imgkit.config(
                                             wkhtmltoimage="C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltoimage.exe"))
                #only if issue key is not in lastseen.
                theBot.send_message(chatid, "Обнаружен новый алерт: "+linky+"!")
                try:
                    theBot.send_photo(chatid, photo=img)
                except Exception as e:
                    if "PHOTO_INVALID_DIMENSIONS" in str(e):
                        Logging(1, "Could not send photo for alert "+issue.key+" using send_photo method. PHOTO_INVALID_DIMENSIONS encountered. Using send_document instead.")
                        theBot.send_document(chatid, img)
                    else:
                        Logging(1, "Exception: "+str(e)+" when sending photo for alert "+issue.key)
            else:
                night = open(os.path.dirname(os.path.abspath(__file__))+'\\nightqueue\\night.txt', 'a+')
                night.write(issue.key+" ")
                imgkit.from_string("<meta charset=\"utf-8\">\n" + issue.fields.description, os.path.dirname(os.path.abspath(__file__))+'\\nightqueue\\'+issue.key+'.png',
                                   config=imgkit.config(
                                       wkhtmltoimage="C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltoimage.exe"))
                nightqueue+=1
        lastseen.add(issue.key)
        lastprocessed = open(os.path.dirname(os.path.abspath(__file__))+'\\lastalert.txt', 'a+')
        lastprocessed.write(issue.key+'\n')
        lastprocessed.close()
    return

#gets text for horoscope
def GetHoroscope(theBot, chatid):
    allowed = ['Водолей', 'Козерог', 'Дева', 'Лев', 'Овен', 'Телец', 'Близнецы', 'Рак', 'Весы', 'Скорпион', 'Стрелец', 'Рыбы']
    h = list()
    try:
        with open(os.path.dirname(os.path.abspath(__file__))+r'\horoscope.txt', 'r', encoding='utf-8') as f:
            h = f.readlines()
        h = [line.rstrip() for line in h]
            
    except Exception as exception:
        Logging(1, 'Exception occured during loading horoscope file: '+str(exception))
    if len(h) == 0:
        Logging(1, 'Empty horoscope file. Either didn\'t request it or permission problems.')
        return
    else:
        if h[0] != datetime.date.today().strftime('%d.%m'):
            theBot.send_message(chatid, 'Ещё не смотрел актуальный гороскоп, извините.')
            Logging(0, 'Lacking up-to-date horoscope.')
        else:
            theBot.send_message(chatid, f'Гороскоп на {h[0]}:')
            for horo in h[1:]:
                for a in allowed:
                    if a in horo[:20]:
                        theBot.send_message(chatid, horo)
    Logging(0, 'Successfully messaged horoscope.')                    
    return

def CraiyonImage(theBot, message, text):
    chatid = message.chat.id
    theBot.reply_to(message, 'Ваш запрос принят в работу! Наспамлю картинок, как только они сгенерируются!')
    getpic = requests.post('https://backend.craiyon.com/generate', json = {
        "prompt": text
    })
    
    try:
        bruh = json.loads(getpic.content)
    except exception as bruh:
        Logging(1, f'Exception during craiyon image loading: {str(bruh)}')
        theBot.reply_to(message, 'К сожалению, что-то пошло не так при получении ответа на запрос! Скорее всего, виновата нейросеть. Попробуйте повторить запрос позже!')
        return

    #print(bruh['images'])
    
    for i in range(len(bruh['images'])):
        bruh['images'][i] = "".join(bruh['images'][i].split('\\n'))

    theBot.reply_to(message, 'Ваш запрос выполнен!')

    for element in bruh['images']:
        sent = False
        image = BytesIO()
        decoded = base64.b64decode(element)
        image.write(decoded)
        image.seek(0)
        while not sent:
            try:    
                theBot.send_photo(chatid, photo=image)
            except Exception as exception:
                Logging(1, f"Exception during sending pics: {str(exception)}")
                if "429" in str(exception):
                    time.sleep(5)
                else:
                    sent = True
            else:
                sent = True
    return

#logging how bot works
def Logging(logflag, message):
    logtype = ('TRACE', 'ERROR')
    if logflag != 0: logflag = 1
    log = open(os.path.dirname(os.path.abspath(__file__))+'\\botlog.txt', 'a+')
    log.write('['+str(datetime.datetime.now())+']['+logtype[logflag]+']'+' '+message+'\n')
    log.close()
    return

def schedule_check(*args):
    cur = threading.current_thread()
    while not args[0].is_set():
        schedule.run_pending()
        time.sleep(1)

class TestException(Exception):
    pass

#main procedure of telegram bot
def mainproc(APIKey, group, groupchatid, chatid, alertchatid, MsgSwitch, killswitch):
 try:
   global lastseen
   try:
      lastprocessed = open(os.path.dirname(os.path.abspath(__file__))+'\\lastalert.txt', 'r')
   except:
       Logging(1, 'Could not open lastalert file, defaulting to None.')
       lastseen = set()
   else:
       lastseen = set(lastprocessed.read().splitlines())
       lastprocessed.close()
   global nightqueue
   try:
       night = open(os.path.dirname(os.path.abspath(__file__)) + '\\nightqueue\\night.txt', 'r')
   except:
       Logging(1, 'Could not open nightqueue file, defaulting to zero.')
       nightqueue = 0
   else:
       nightlist = night.read().split(" ")
       if len(nightlist[-1]) == 0: del nightlist[-1]
       nightqueue = len(nightlist)
       night.close()
   #debug threading
   for thread in threading.enumerate():
     print(thread.name)

   theBot = telebot.TeleBot(APIKey)
   theBot.send_message(chatid, 'Awakening...')
   Logging(0, 'Launching')

   if MsgSwitch != 0:
       schedule.every(1).minutes.do(GetAlert, theBot, alertchatid) #alertchatid when done !!!
       sub_stop = threading.Event()
       sub = threading.Thread(target=schedule_check, args=(sub_stop,)).start()
       """ #old scheduler, dont really need
       schedule.every().sunday.at("15:00").do(DutyInform, theBot, group, groupchatid, chatid)
       sub_stop = threading.Event()
       sub = threading.Thread(target=schedule_check, args=(sub_stop,)).start()
       """

   @theBot.message_handler(commands=['Exit', 'exit'])
   def Exiting(message):
      if message.chat.id not in {groupchatid, chatid, alertchatid}:
            theBot.send_message(message.chat.id, "Не доверяю данному чату с этой командой.")
            return
      theBot.send_message(message.chat.id, "Shutting down...")
      Logging(0, 'Shutting down')
      if MsgSwitch != 0:
          schedule.clear()
          sub_stop.set()
      theBot.stop_polling() #this is basically exit() for telegrambotAPI
      killswitch[0] = False
      return

   @theBot.message_handler(commands=['дежурство', 'Дежурство'])
   def SendDuty(message):
      if message.chat.id not in {groupchatid, chatid, alertchatid}:
            theBot.send_message(message.chat.id, "Не доверяю данному чату с этой командой.")
            return
      FIOCheck = list()
      customrange = -1
      if len(message.text) > len("/дежурство")+1:
          temp = message.text[len("/дежурство")+1:len(message.text)]
          i = 0
          for x in range(len(temp)):
              if temp[x] == '|':
                  try:
                      customrange = int(temp[x+1:len(temp)].strip())
                  except:
                      Logging(1, "could not convert custom range to int. String was: "+str(temp[x+1:len(temp)].strip()))
                  temp = temp[:x]
                  break
          while i < len(temp):
              if temp[i] == ' ':
                  if i > 0: FIOCheck.append(temp[:i])
                  temp = temp[i+1:len(temp)]
                  i = 0
              else:
                  i+=1
          if len(temp) > 0: FIOCheck.append(temp)
      if len(group) > 0: grouptext = 'группе '+group
      else: grouptext = 'всем группам'
      if customrange != -1: period = ' следующее количество дней: '+str(customrange)
      else: period = ' ближайшую неделю'
      if len(FIOCheck) > 0: ifFIO = " (дополнительно фильтрую по ФИО сотрудника)"
      else: ifFIO = ""
      theBot.send_message(message.chat.id, "График дежурств по "+grouptext+" на"+str(period)+str(ifFIO)+": ")
      Logging(0, 'Duty table request at channelID '+str(message.chat.id))
      Logging(0, 'Preparing duty table')
      try:
        img = GetData(address, group, FIOCheck, customrange)
      except Exception as exception:
        Logging(1, 'Exception occured: '+str(exception))
        theBot.send_message(message.chat.id, "Не смог получить график. Либо не найдена нужная таблица, либо данных нет или они в некорректном формате.")
      else:
        img.seek(0)
        theBot.send_photo(message.chat.id, photo=img)
        Logging(0, 'Sent duty table')

   @theBot.message_handler(commands=['полезное', 'Полезное'])
   def UsefulLinks(message):
       if message.chat.id not in {groupchatid, chatid, alertchatid}:
           theBot.send_message(message.chat.id, "Не доверяю данному чату с этой командой.")
           return
       linkspath = os.path.join(os.path.dirname(os.path.realpath(__file__)), "UsefulLinks.xlsx")
       Logging(0, "UsefulLinks request at chat id "+str(message.chat.id))
       try:
         linklist = UsefulLinksList(linkspath)
         for i in range(len(linklist)):
             linklist[i] = ' - '.join(linklist[i])
         theBot.send_message(message.chat.id, "Полезные ссылки:"+
                             "\n\n"+'\n\n'.join(linklist))
       except Exception as exception:
           Logging(1, 'Failed to get useful links with exception: ' + str(exception))
           theBot.send_message(message.chat.id, "Не удалось получить список полезных ссылок.")
       return
       
   @theBot.message_handler(commands=['Horoscope', 'horoscope', 'Гороскоп', 'гороскоп', 'каксегодняжить', 'Каксегодняжить'])
   def Horoscope(message):
       Logging(0, 'Requested horoscope')
       GetHoroscope(theBot, message.chat.id)
       return
    
   @theBot.message_handler(commands=['Memepic', 'memepic', 'Craiyon', 'craiyon', 'картинка', 'Картинка'])
   def Craiyon(message):
       i = 0
       while i<len(message.text) and message.text[i] != ' ':
            i+=1
       if i < len(message.text):
            prompt = message.text[i:]
       else:
            prompt = 'Здравствуйте'
       Logging(0, 'Requested craiyon image')
       CraiyonImage(theBot, message, prompt)
       return

   @theBot.message_handler(commands=['Help', 'help', 'Хэлп', 'хэлп', 'помощь', 'Помощь', 'Команды', 'команды'])
   def Help(message):
       Logging(0, 'Commands list message')
       theBot.send_message(message.chat.id, "Список команд:\n\n/дежурство <опционально текст для фильтрации по ФИО> | <опционально кол-во дней>: запросить график дежурство на указанное количество дней по указанному ФИО сотрудника;"+
                                            "\n\n/полезное: запросить полезные ссылки;"+
                                            "\n\nздравствуйте - поздороваться;"+
                                            "\n\n/horoscope - запросить гороскоп на сегодня, если он есть.")
       return

   @theBot.message_handler(commands=['coolstory_old9999'])
   def Coolstory_old(message):
       Logging(0, 'Balaboba message')
       balabob = {
           'Content-Type': 'application/json',
           'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 11_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.1 Safari/605.1.15',
           'Origin': 'https://yandex.ru',
           'Referer': 'https://yandex.ru/',
       }
       if len(message.text) > len("/coolstory"):
           basis = message.text[len("/coolstory"):len(message.text)].strip()
       else:
           basis = "Группа ОФД"
       print(basis)
       payload = {"query": basis, "intro": 6, "filter": 1}
       params = json.dumps(payload).encode('utf8')
       req = urllib.request.Request('https://zeapi.yandex.net/lab/api/yalm/text3', data=params, headers=balabob)

       response = urllib.request.urlopen(req)
       if response.code == 200:
           temp = response.read().decode('unicode-escape')#.replace('"','\\"')
           print(temp)
           i = 0
           while temp[i:i+len("bad_query")] != "bad_query": i+=1
           i = i+len("bad_query")+2
           print(temp[i])
           if temp[i] == '1':
               superstring = "Цензура не пропустила этот запрос. Во всем виноват яндекс!"
           else:
                while temp[i:i+len("text")] != "text": i+=1
                i = i+len("text")+3
                print(temp[i:len(temp)-3])
                superstring = temp[i:len(temp)-3]
                #superstring = json.loads(temp, strict=False)["text"]
       else:
           superstring = "Почему-то не смог получить интересную историю. Во всем виноват яндекс!"
       if len(superstring) == 0: superstring = "Вернулся пустой текст. Во всем виноват яндекс!"
       theBot.reply_to(message, superstring)
       return
    
   @theBot.message_handler(commands=['coolstory', 'Coolstory'])
   def Coolstory(message):
        Logging(0, 'Porfirievich message')
        random.seed()
        if len(message.text) > len("/coolstory"):
           basis = message.text[len("/coolstory"):len(message.text)].strip()  
        else:
           prompts = ["Присаживайся и послушай: ", "Увлекательная история: ", "Здравствуйте! "]
           basis = prompts[random.randrange(4)]
        params = {"prompt" : basis, "length" : random.randrange(75)+75}
        getstory = requests.post("https://pelevin.gpt.dobro.ai/generate/", json=params)
        
        if getstory.status_code == 200:
            theBot.reply_to(message, basis+json.loads(getstory.content)['replies'][0])
            Logging(0, 'Porfirievich message sent.')
        else:
            theBot.reply_to(message, "Почему-то не смог найти историю, извините.")
            Logging(1, f'Error {getstory.status_code} when trying to get text from neural network.')
        return

   @theBot.message_handler(commands=['alerttttttt']) #old deprecated command dont use for now. lastseen is not set. GetAlert returns nothing instead of issue key
   def NewAlert(message):
       global lastseen
       Logging(0, 'Alert invoker')
       lastseen = GetAlert(theBot, message.chat.id)
       return

   @theBot.message_handler(content_types = ["new_chat_members"])
   def GreetNew(message):
      Logging(0, 'Welcoming new user')
      theBot.reply_to(message, "Добро пожаловать!")
      return

   theBot.infinity_polling(timeout=25, long_polling_timeout = 5)
 except Exception as exception:
   Logging(1, 'Failed during idling with exception: '+str(exception))
   Logging(0, 'Resuming work in 60 seconds')
   if MsgSwitch != 0:
       schedule.clear()
       sub_stop.set()
   killswitch[0] = True
   theBot.stop_polling()
   time.sleep(60)
   Logging(0, 'Trying to resume. Killswitch is '+str(killswitch[0]))
   #mainproc(APIKey, group, groupchatid, chatid, MsgSwitch)

if __name__ == "__main__":
 #the empty string is so you can set ini file to 3 and get data for all groups:
 #WeeklyMsg - to turn on weekly message. 0 = off, else on
 groups = ('ЭДО ГКО', 'ЭДО', 'УЦ', 'ОФД', '')
 config = configparser.ConfigParser()
 try:
   config.read(os.path.dirname(os.path.abspath(__file__))+'\\botconfig.ini')
 except:
   Logging(1, 'Could not find settings file.')
   exit()

 if int(config['Primary']['GroupFlag']) not in {0, 1, 2, 3, 4}: config['Primary']['GroupFlag'] = 4

 killswitch = [True]
 while killswitch[0]:
    mainproc(config['Primary']['APIKey'],
          groups[int(config['Primary']['GroupFlag'])],
          int(config['Primary']['Groupchat']),
          int(config['Primary']['Testchat']),
          int(config['Primary']['Alertchat']),
              int(config['Primary']['AlertInform']), killswitch)
 log = open(os.path.dirname(os.path.abspath(__file__))+'\\botlog.txt', 'a')
 log.write('\n')
 log.close()